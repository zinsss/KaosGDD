import io
import pathlib
import sys
import tempfile
import unittest
from unittest import mock


APP_ROOT = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(APP_ROOT))

from services.ledger import import_xlsx, service


class LedgerValidationTests(unittest.TestCase):
    def test_all_legacy_categories_calculate_expected_deltas(self):
        cases = {
            "계좌 수입": (100, 0, 0),
            "계좌 지출": (-100, 0, 0),
            "현금 인출": (-100, 100, 0),
            "계좌 입금": (100, -100, 0),
            "상품권 구입 - 계좌": (-100, 0, 100),
            "현금 수입": (0, 100, 0),
            "현금 지출": (0, -100, 0),
            "상품권 구입 - 현금": (0, -100, 100),
            "상품권 사용": (0, 0, -100),
        }
        for category, expected in cases.items():
            with self.subTest(category=category):
                self.assertEqual(service.deltas_for(category, 100)[2], expected)

    def test_running_balances_preserve_entry_order(self):
        entries = [
            {"accountDelta": 0, "cashDelta": 1000, "giftDelta": 0},
            {"accountDelta": 0, "cashDelta": -200, "giftDelta": 200},
            {"accountDelta": 0, "cashDelta": 0, "giftDelta": -50},
        ]
        balanced, totals = service._with_balances(entries)
        self.assertEqual(balanced[1]["cash"], 800)
        self.assertEqual(totals, {"account": 0, "cash": 800, "gift": 150})


class LegacyWorkbookTests(unittest.TestCase):
    def build_workbook(self, path, broken=False):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "마야병원 의사회비 장부"
        sheet.append(list(import_xlsx.EXPECTED_HEADERS))
        sheet.append(["2023.01.01", "인수인계", None, "시작", 0, 1000, 0])
        sheet.append(["2023.01.02", "상품권 구입 - 현금", 200, "구입", 0, 800, 200])
        sheet.append(["2023.01.03", "상품권 사용", 50, "사용", 0, 800, 999 if broken else 150])
        workbook.save(path)

    def test_source_import_reconciles_every_cached_balance(self):
        with tempfile.TemporaryDirectory() as directory:
            path = pathlib.Path(directory) / "source.xlsx"
            self.build_workbook(path)
            source = import_xlsx.load_source(path)
        self.assertEqual(len(source["rows"]), 3)
        self.assertTrue(source["rows"][0]["locked"])
        self.assertEqual(source["rows"][1]["source_row"], 4)
        self.assertEqual(source["finalBalances"], {"account": 0, "cash": 800, "gift": 150})

    def test_source_import_rejects_a_balance_mismatch(self):
        with tempfile.TemporaryDirectory() as directory:
            path = pathlib.Path(directory) / "broken.xlsx"
            self.build_workbook(path, broken=True)
            with self.assertRaisesRegex(ValueError, "source_balance_mismatch_row_5"):
                import_xlsx.load_source(path)


class LedgerExportTests(unittest.TestCase):
    @mock.patch.object(service, "_audit_rows", return_value=[])
    @mock.patch.object(service, "list_ledger")
    def test_export_is_a_readable_multisheet_workbook(self, list_ledger, _audit_rows):
        list_ledger.return_value = {
            "ok": True,
            "entryCount": 1,
            "balances": {"account": 0, "cash": 1000, "gift": 0},
            "entries": [
                {
                    "date": "2023-01-01",
                    "category": "인수인계",
                    "amount": None,
                    "details": "시작",
                    "account": 0,
                    "cash": 1000,
                    "gift": 0,
                    "accountDelta": 0,
                    "cashDelta": 1000,
                    "giftDelta": 0,
                }
            ],
        }
        data = service.workbook_bytes()

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), data_only=True)
        self.assertEqual(workbook.sheetnames, ["거래내역", "월별요약", "변경기록", "정보"])
        self.assertEqual(workbook["거래내역"]["F2"].value, 1000)
        workbook.close()

import argparse
import hashlib
import json
from datetime import date, datetime
from pathlib import Path

from models.database import connect
from services.ledger import service


EXPECTED_HEADERS = ("날짜", "사용 구분", "금액", "상세 내용", "계좌", "현금", "상품권")


def source_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    normalized = str(value or "").strip().replace(".", "-").rstrip("-")
    return date.fromisoformat(normalized)


def source_integer(value, allow_empty=False):
    if value in (None, "") and allow_empty:
        return None
    if isinstance(value, bool):
        raise ValueError("invalid_source_amount")
    number = int(value)
    if float(value) != number:
        raise ValueError("non_integer_source_amount")
    return number


def load_source(path):
    from openpyxl import load_workbook

    path = Path(path)
    checksum = hashlib.sha256(path.read_bytes()).hexdigest()
    formula_book = load_workbook(path, data_only=False, read_only=False)
    value_book = load_workbook(path, data_only=True, read_only=False)
    try:
        formula_sheet = formula_book.active
        value_sheet = value_book.active
        headers = tuple(str(value_sheet.cell(2, column).value or "").strip() for column in range(1, 8))
        if headers != EXPECTED_HEADERS:
            raise ValueError("unexpected_ledger_headers")
        rows = []
        account = cash = gift = 0
        for row_number in range(3, value_sheet.max_row + 1):
            values = [value_sheet.cell(row_number, column).value for column in range(1, 8)]
            if values[0] in (None, ""):
                continue
            entry_date = source_date(values[0])
            category = str(values[1] or "").strip()
            amount = source_integer(values[2], allow_empty=True)
            details = str(values[3] or "").strip()
            cached = tuple(source_integer(value) for value in values[4:7])
            if not rows:
                if category != "인수인계":
                    raise ValueError("missing_opening_handover")
                deltas = cached
                locked = True
            else:
                normalized_category, normalized_amount, deltas = service.deltas_for(category, amount)
                category = normalized_category
                amount = normalized_amount
                locked = False
            account += deltas[0]
            cash += deltas[1]
            gift += deltas[2]
            if (account, cash, gift) != cached:
                raise ValueError(f"source_balance_mismatch_row_{row_number}")
            rows.append(
                {
                    "id": f"legacy-{checksum[:12]}-{row_number}",
                    "sort_order": len(rows) * 1000 + 1000,
                    "entry_date": entry_date,
                    "category": category,
                    "amount": amount,
                    "details": details,
                    "deltas": deltas,
                    "source_row": row_number,
                    "locked": locked,
                    "cached": cached,
                    "formula": tuple(formula_sheet.cell(row_number, column).value for column in range(5, 8)),
                }
            )
        if not rows:
            raise ValueError("empty_ledger_source")
        return {
            "path": str(path),
            "sha256": checksum,
            "sheet": value_sheet.title,
            "rows": rows,
            "finalBalances": {"account": account, "cash": cash, "gift": gift},
        }
    finally:
        formula_book.close()
        value_book.close()


def import_source(path, actor="migration", dry_run=False):
    source = load_source(path)
    report = {
        "ok": True,
        "dryRun": bool(dry_run),
        "source": source["path"],
        "sha256": source["sha256"],
        "sheet": source["sheet"],
        "entryCount": len(source["rows"]),
        "finalBalances": source["finalBalances"],
    }
    if dry_run:
        return report
    with connect() as connection:
        with connection.transaction():
            connection.execute("SELECT pg_advisory_xact_lock(hashtext('family-ledger'))")
            if connection.execute("SELECT count(*) FROM family_ledger_entries").fetchone()[0]:
                raise ValueError("ledger_not_empty")
            for item in source["rows"]:
                row = connection.execute(
                    """
                    INSERT INTO family_ledger_entries (
                        id, sort_order, entry_date, category, amount, details,
                        account_delta, cash_delta, gift_delta, source_row,
                        source_checksum, locked, created_by, updated_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id, sort_order, entry_date, category, amount, details,
                              account_delta, cash_delta, gift_delta, source_row, source_checksum,
                              locked, revision, created_by, updated_by, created_at, updated_at
                    """,
                    (
                        item["id"], item["sort_order"], item["entry_date"], item["category"], item["amount"],
                        item["details"], *item["deltas"], item["source_row"], source["sha256"], item["locked"],
                        actor, actor,
                    ),
                ).fetchone()
                imported = service._entry_from_row(row)
                service._audit(connection, item["id"], "import", actor, None, imported)
    service.write_backup("latest")
    return report


def main():
    parser = argparse.ArgumentParser(description="Import the legacy family ledger XLSX into an empty Brain ledger.")
    parser.add_argument("path")
    parser.add_argument("--actor", default="migration")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    print(json.dumps(import_source(args.path, actor=args.actor, dry_run=args.dry_run), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
